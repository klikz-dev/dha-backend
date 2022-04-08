from django.core.management.base import BaseCommand

import os
import openpyxl
import re

from membersuite.models import Membership

BASEDIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REG_EMAIL = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'


class Command(BaseCommand):
    help = 'Write Reports data from excel sheet'

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "main" in options['functions']:
            self.main()

    def main(self):
        Membership.objects.all().delete()

        files = ['DHMW20-Reports.xlsx', 'DHMW21-Reports.xlsx']

        for file in files:
            wb = openpyxl.load_workbook(BASEDIR + '/files/{}'.format(file))
            sh = wb.active

            for i in range(2, sh.max_row):
                email = str(sh.cell(row=i, column=5).value).strip()

                if not re.fullmatch(REG_EMAIL, email):
                    continue

                try:
                    Membership.objects.get(email=email)
                    continue
                except Membership.DoesNotExist:
                    pass

                membership = 'provider'

                acute = str(sh.cell(row=i, column=16).value).strip()
                if acute == 'N/A':
                    acute = ''

                ambulatory = str(sh.cell(row=i, column=17).value).strip()
                if ambulatory == 'N/A':
                    ambulatory = ''

                ltpac = str(sh.cell(row=i, column=18).value).strip()
                if ltpac == 'N/A':
                    ltpac = ''

                organization_name = str(sh.cell(row=i, column=2).value).strip()

                Membership.objects.create(
                    email=email,
                    membership=membership,
                    acute=acute,
                    ambulatory=ambulatory,
                    ltpac=ltpac,
                    organization_name=organization_name
                )

                print("Success to get membership details for email: {}".format(email))
